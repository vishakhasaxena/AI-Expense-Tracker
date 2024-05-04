from flask import Flask, render_template, request, redirect, url_for
from datetime import datetime
import openpyxl

app = Flask(__name__)

def update_expense_sheet(category, amount):
    workbook = openpyxl.load_workbook('expenses.xlsx')
    sheet = workbook.active
    sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), category, amount])
    workbook.save('expenses.xlsx')

def analyze_expenses():
    threshold = 0.7
    suggestions = []
    workbook = openpyxl.load_workbook('expenses.xlsx')
    sheet = workbook.active
    expenses = [row[2].value for row in sheet.iter_rows(min_row = 2,max_col = 3,values_only = True)]
    total_expenses = sum(expenses)
    remaining_budget = 1000 - total_expenses
    if remaining_budget < 0:
        suggestions.append("You have exceeded your budget!")
    for expense in expenses:
        if expense > threshold * total_expenses:
            suggestions.append(f"Consider minimizing expenses in the category of ${expense}")
    return suggestions, remaining_budget

def home():
    suggestions, remaining_budget = analyze_expenses()
    return render_template('index.html', suggestions = suggestions, remaining_budget = remaining_budget)

def submit_expense():
    category = request.form['category']
    amount = float(request.form['amount'])
    update_expense_sheet(category, amount)
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug = True)

